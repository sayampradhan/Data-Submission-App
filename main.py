# Import necessary modules
from tkinter import *
from tkinter.ttk import Combobox
import tkinter.messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

# Check if the 'data_entry.xlsx' file exists, and create it with headers if it doesn't
file = input("File Name (without .xlsx) : ")
file_path = pathlib.Path(file + ".xlsx")
if file_path.exists():
    print("File already exists.")
else:
    # Create a new workbook and add headers
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save('data_entry.xlsx')


# Placeholder functions for submit and clear actions
def submit():
    # Retrieve form values
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Load the existing workbook or create a new one
    xlsx_file = openpyxl.load_workbook('data_entry.xlsx')
    sheet = xlsx_file.active

    # Find the next available row and add the data
    next_row = sheet.max_row + 1
    sheet.cell(column=1, row=next_row, value=name)
    sheet.cell(column=2, row=next_row, value=contact)
    sheet.cell(column=3, row=next_row, value=age)
    sheet.cell(column=4, row=next_row, value=gender)
    sheet.cell(column=5, row=next_row, value=address)

    # Save the workbook
    xlsx_file.save('data_entry.xlsx')

    # Display a message to confirm the submission
    tkinter.messagebox.showinfo('Info', "Data has been submitted successfully.")

    # Clear form fields after submission
    clear()


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)


# Create the main Tkinter window
root = Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#FFFBC8")

# Set the application icon
icon_image = PhotoImage(file='logo.png')
root.iconphoto(False, icon_image)

# Create and place a heading label
Label(root, text="Please fill out this Entry form:", font="arial 13", bg="#FFFBC8", fg="#48494B").place(x=20, y=20)

# Labels for form fields
Label(root, text='Name', font=23, bg="#FFFBC8", fg="#48494B").place(x=50, y=100)
Label(root, text='Contact No.', font=23, bg="#FFFBC8", fg="#48494B").place(x=50, y=150)
Label(root, text='Age', font=23, bg="#FFFBC8", fg="#48494B").place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#FFFBC8", fg="#48494B").place(x=390, y=200)
Label(root, text='Address', font=23, bg="#FFFBC8", fg="#48494B").place(x=50, y=250)

# Entry widgets
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=40, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=40, bd=2, font=20)
AgeEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)

# Gender Combobox
gender_combobox = Combobox(root, values=['Male', 'Female'], font='arial 14', state='readonly', width=14)
gender_combobox.place(x=470, y=200)
gender_combobox.set('Male')

addressEntry = Text(root, width=55, height=4, bd=4)

# Place entry widgets
nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
AgeEntry.place(x=200, y=200)
addressEntry.place(x=200, y=250)

# Buttons for submit, clear, and exit
Button(root, text="SUBMIT", bg="#222021", fg="white", width=15, height=1, command=submit).place(x=200, y=350)
Button(root, text="CLEAR", bg="#222021", fg="white", width=15, height=1, command=clear).place(x=340, y=350)
Button(root, text="EXIT", bg="#222021", fg="white", width=15, height=1, command=root.destroy).place(x=480, y=350)

# Start the main loop
root.mainloop()
